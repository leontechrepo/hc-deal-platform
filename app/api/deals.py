import re
import uuid
from datetime import date, datetime, timezone
from io import BytesIO
from typing import Optional

import openpyxl
from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from pydantic import BaseModel
from sqlalchemy import delete, select, text, desc
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.auth import get_actor_name, require_auth
from app.core.config import settings
from app.db.activity import log_activity
from app.db.approvals import log_approval
from app.db.models import Deal, DealDocument, DealUpdateLog, EmailScanLog
from app.db.models.companies import Company
from app.db.portfolio import ensure_portfolio_position
from app.db.session import get_db
from app.domain.pipeline_stage import (
    PIPELINE_STAGES,
    STATUSES,
    UNDERWRITING_FIELDS,
    is_underwriting_locked,
)
from app.storage import documents as storage

router = APIRouter(prefix="/api", dependencies=[Depends(require_auth)])

# Legacy fields (kept editable — the old Dashboard's inline-edit still targets
# these) plus every new structural/financial/covenant field from migration 005
# and the new pipeline_stage/status columns from migration 004.
EDITABLE_FIELDS = {
    "company_name", "location", "stage", "sector_primary", "sector_full",
    "subsector", "deal_size_m", "security", "uop", "source", "commentary",
    "reasons_for_passing", "bucket",
    "pipeline_stage", "status",
    "state", "next_action", "sourcing_date", "contact_name", "contact_role",
    "nda_date", "nda_status", "tenor_months", "amortization", "oid_pct",
    "sofr_floor_pct", "call_protection", "maturity_date", "total_leverage",
    "spread_bps", "base_rate", "sofr_rate", "all_in_rate", "hold_amount_m",
    "revenue_growth_pct", "ebitda_growth_pct", "capex_m", "fcf_m", "dscr",
    "fccr", "interest_coverage", "max_leverage_covenant", "min_fccr_covenant",
    "capex_limit_covenant_m", "employees", "locations_count", "year_founded",
    "risk_score", "ltm_revenue_m", "ltm_ebitda_m", "ebitda_margin",
}

_DATE_FIELDS = {"target_close", "sourcing_date", "nda_date", "maturity_date"}
_NUMERIC_FIELDS = {
    "deal_size_m", "total_funded_m", "ltm_revenue_m", "ltm_ebitda_m", "ebitda_margin",
    "oid_pct", "sofr_floor_pct", "total_leverage", "sofr_rate", "all_in_rate",
    "hold_amount_m", "revenue_growth_pct", "ebitda_growth_pct", "capex_m", "fcf_m",
    "dscr", "fccr", "interest_coverage", "max_leverage_covenant", "min_fccr_covenant",
    "capex_limit_covenant_m", "risk_score",
}
_INT_FIELDS = {"tenor_months", "spread_bps", "employees", "locations_count", "year_founded"}


def _coerce_field_value(field: str, value):
    if value is None or value == "":
        return None
    if field in _DATE_FIELDS:
        return value if isinstance(value, date) else date.fromisoformat(str(value))
    if field in _NUMERIC_FIELDS:
        return float(value)
    if field in _INT_FIELDS:
        return int(value)
    return value


class PatchRequest(BaseModel):
    field: str
    value: str | float | int | None = None


@router.patch("/deals/{deal_id}")
async def patch_deal(
    deal_id: uuid.UUID,
    body: PatchRequest,
    db: AsyncSession = Depends(get_db),
    auth: dict = Depends(require_auth),
):
    if body.field not in EDITABLE_FIELDS:
        raise HTTPException(status_code=400, detail=f"Field '{body.field}' is not editable")

    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    if body.field in UNDERWRITING_FIELDS and is_underwriting_locked(deal.pipeline_stage):
        raise HTTPException(
            status_code=409,
            detail=(
                f"'{body.field}' is locked — underwriting fields become read-only once a "
                "deal reaches loi_signed or later. Contact an admin to unlock."
            ),
        )
    if body.field == "pipeline_stage" and body.value not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid pipeline_stage: {body.value!r}")
    if body.field == "status" and body.value not in STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {body.value!r}")

    old_value_raw = getattr(deal, body.field)

    try:
        coerced = _coerce_field_value(body.field, body.value)
    except (TypeError, ValueError) as exc:
        raise HTTPException(status_code=400, detail=f"Invalid value for '{body.field}': {exc}")

    # Compare numerically (not by string) so an unchanged numeric field isn't
    # spuriously flagged as "changed" just because SQLAlchemy round-trips it
    # as e.g. Decimal("20.00") while the coerced input is float 20.0.
    if body.field in _NUMERIC_FIELDS:
        value_changed = (float(old_value_raw) if old_value_raw is not None else None) != coerced
    else:
        value_changed = old_value_raw != coerced

    old_value = str(old_value_raw) if old_value_raw is not None else None
    new_value = str(coerced) if coerced is not None else None

    setattr(deal, body.field, coerced)
    deal.updated_by = "manual_edit"
    deal.updated_at = datetime.now(timezone.utc)

    if body.field == "pipeline_stage" and coerced == "portfolio_monitoring":
        await ensure_portfolio_position(deal, db)

    # Skip the log/activity entries when the "edit" didn't actually change the value
    # (e.g. re-saving a field unchanged) — otherwise the Logs page shows a colored
    # diff between two identical values, which reads as a change that never happened.
    if value_changed:
        log = DealUpdateLog(
            deal_id=deal_id,
            field_changed=body.field,
            old_value=old_value,
            new_value=new_value,
            source="manual_edit",
        )
        db.add(log)

        activity_type = "stage_change" if body.field in ("pipeline_stage", "stage", "bucket") else \
            "status_change" if body.field == "status" else "system"
        await log_activity(
            db, deal_id, get_actor_name(auth), activity_type,
            f"{body.field} changed from {old_value!r} to {coerced!r}",
        )
        if body.field in ("pipeline_stage", "status"):
            await log_approval(db, deal_id, str(coerced), get_actor_name(auth))

    return {"ok": True, "deal_id": str(deal_id), "field": body.field, "value": coerced}


class DealUpdateRequest(BaseModel):
    """Bulk-edit counterpart to PatchRequest — every field in EDITABLE_FIELDS,
    all optional so a partial "Edit Deal" form save only touches changed fields."""

    company_name: Optional[str] = None
    location: Optional[str] = None
    stage: Optional[str] = None
    sector_primary: Optional[str] = None
    sector_full: Optional[str] = None
    subsector: Optional[str] = None
    deal_size_m: Optional[float] = None
    security: Optional[str] = None
    uop: Optional[str] = None
    source: Optional[str] = None
    commentary: Optional[str] = None
    reasons_for_passing: Optional[str] = None
    bucket: Optional[str] = None
    pipeline_stage: Optional[str] = None
    status: Optional[str] = None
    state: Optional[str] = None
    next_action: Optional[str] = None
    sourcing_date: Optional[str] = None
    contact_name: Optional[str] = None
    contact_role: Optional[str] = None
    nda_date: Optional[str] = None
    nda_status: Optional[str] = None
    tenor_months: Optional[int] = None
    amortization: Optional[str] = None
    oid_pct: Optional[float] = None
    sofr_floor_pct: Optional[float] = None
    call_protection: Optional[str] = None
    maturity_date: Optional[str] = None
    total_leverage: Optional[float] = None
    spread_bps: Optional[int] = None
    base_rate: Optional[str] = None
    sofr_rate: Optional[float] = None
    all_in_rate: Optional[float] = None
    hold_amount_m: Optional[float] = None
    revenue_growth_pct: Optional[float] = None
    ebitda_growth_pct: Optional[float] = None
    capex_m: Optional[float] = None
    fcf_m: Optional[float] = None
    dscr: Optional[float] = None
    fccr: Optional[float] = None
    interest_coverage: Optional[float] = None
    max_leverage_covenant: Optional[float] = None
    min_fccr_covenant: Optional[float] = None
    capex_limit_covenant_m: Optional[float] = None
    employees: Optional[int] = None
    locations_count: Optional[int] = None
    year_founded: Optional[int] = None
    risk_score: Optional[float] = None
    ltm_revenue_m: Optional[float] = None
    ltm_ebitda_m: Optional[float] = None
    ebitda_margin: Optional[float] = None


@router.put("/deals/{deal_id}")
async def update_deal(
    deal_id: uuid.UUID,
    body: DealUpdateRequest,
    db: AsyncSession = Depends(get_db),
    auth: dict = Depends(require_auth),
):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    updates = body.model_dump(exclude_unset=True)
    if not updates:
        return {"ok": True, "deal_id": str(deal_id), "updated_fields": [], "deal": _deal_to_dict(deal)}

    unknown = set(updates) - EDITABLE_FIELDS
    if unknown:
        raise HTTPException(status_code=400, detail=f"Fields not editable: {sorted(unknown)}")

    if "pipeline_stage" in updates and updates["pipeline_stage"] not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid pipeline_stage: {updates['pipeline_stage']!r}")
    if "status" in updates and updates["status"] not in STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {updates['status']!r}")

    coerced_updates: dict = {}
    for field, raw_value in updates.items():
        try:
            coerced_updates[field] = _coerce_field_value(field, raw_value)
        except (TypeError, ValueError) as exc:
            raise HTTPException(status_code=400, detail=f"Invalid value for '{field}': {exc}")

    # Only fields whose coerced value actually differs from the deal's
    # current one count as "changed" — an edit-modal save that resubmits the
    # whole record (locked underwriting fields included, unchanged) must be a
    # no-op for those fields, not a lock violation.
    changed_fields: dict = {}
    for field, coerced in coerced_updates.items():
        old_value_raw = getattr(deal, field)
        if field in _NUMERIC_FIELDS:
            value_changed = (float(old_value_raw) if old_value_raw is not None else None) != coerced
        else:
            value_changed = old_value_raw != coerced
        if value_changed:
            changed_fields[field] = coerced

    if not changed_fields:
        return {"ok": True, "deal_id": str(deal_id), "updated_fields": [], "deal": _deal_to_dict(deal)}

    # Locked against the stage this request would leave the deal in
    # (applying any pipeline_stage change from this same request), not the
    # stage before it — otherwise a single call could advance pipeline_stage
    # to loi_signed and edit a locked field in that same request.
    resulting_stage = changed_fields.get("pipeline_stage", deal.pipeline_stage)
    locked = is_underwriting_locked(resulting_stage)
    lock_violations = sorted(f for f in changed_fields if f in UNDERWRITING_FIELDS and locked)
    if lock_violations:
        raise HTTPException(
            status_code=409,
            detail=(
                f"Fields locked — underwriting fields become read-only once a deal "
                f"reaches loi_signed or later: {lock_violations}"
            ),
        )

    for field, coerced in changed_fields.items():
        old_value_raw = getattr(deal, field)
        old_value = str(old_value_raw) if old_value_raw is not None else None
        new_value = str(coerced) if coerced is not None else None
        setattr(deal, field, coerced)
        db.add(DealUpdateLog(
            deal_id=deal_id, field_changed=field, old_value=old_value, new_value=new_value,
            source="manual_edit",
        ))

    changed_field_names = list(changed_fields.keys())

    deal.updated_by = "manual_edit"
    deal.updated_at = datetime.now(timezone.utc)

    if "pipeline_stage" in changed_fields and deal.pipeline_stage == "portfolio_monitoring":
        await ensure_portfolio_position(deal, db)

    # One combined activity entry (avoids flooding the Activity tab on a
    # multi-field form save), plus a dedicated stage/status entry — matches
    # today's narrative style where transitions get their own line.
    if any(f in ("pipeline_stage", "stage", "bucket") for f in changed_field_names):
        await log_activity(
            db, deal_id, get_actor_name(auth), "stage_change",
            f"Pipeline stage/bucket updated to {deal.pipeline_stage!r}/{deal.bucket!r}",
        )
    if "status" in changed_field_names:
        await log_activity(db, deal_id, get_actor_name(auth), "status_change", f"Status changed to {deal.status!r}")
    if "pipeline_stage" in changed_field_names:
        await log_approval(db, deal_id, deal.pipeline_stage, get_actor_name(auth))
    if "status" in changed_field_names:
        await log_approval(db, deal_id, deal.status, get_actor_name(auth))
    other_fields = [f for f in changed_field_names if f not in ("pipeline_stage", "stage", "bucket", "status")]
    if other_fields:
        summary = ", ".join(other_fields[:5]) + (f" (+{len(other_fields) - 5} more)" if len(other_fields) > 5 else "")
        await log_activity(db, deal_id, get_actor_name(auth), "system", f"Deal updated — fields changed: {summary}")

    return {"ok": True, "deal_id": str(deal_id), "updated_fields": changed_field_names, "deal": _deal_to_dict(deal)}


@router.delete("/deals/{deal_id}")
async def delete_deal(
    deal_id: uuid.UUID,
    db: AsyncSession = Depends(get_db),
    auth: dict = Depends(require_auth),
):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    company_name = deal.company_name

    # Blob storage isn't covered by the DB's ON DELETE CASCADE — collect keys
    # before the row disappears out from under us.
    doc_rows = await db.execute(
        select(DealDocument.storage_key).where(
            DealDocument.deal_id == deal_id, DealDocument.storage_key.is_not(None)
        )
    )
    storage_keys = [k for (k,) in doc_rows.all()]

    # Single statement — Postgres's own ON DELETE CASCADE/SET NULL cascades to
    # deal_update_log, deal_activity, deal_notes, deal_documents,
    # deal_timeline_workstreams (+tasks), portfolio_positions (+monitoring
    # tests), and pending_suggestions; nulls out email_scan_log.matched_deal_id
    # and chat_sessions.deal_id.
    await db.execute(delete(Deal).where(Deal.id == deal_id))

    # Commit the deletion before touching storage: if the object delete
    # succeeded but the transaction then failed or rolled back, the deal and
    # document rows would remain while their only stored files are already
    # gone. Committing first means a failed storage delete only leaves a
    # harmless orphaned object, never a dangling reference to a deleted one
    # (same reasoning as delete_document's storage cleanup).
    await db.commit()

    # Best-effort blob cleanup — a storage failure here must never roll back
    # the deal deletion; an orphaned blob is an acceptable failure mode.
    if storage_keys and settings.storage_configured:
        for key in storage_keys:
            try:
                storage.delete_object(key)
            except Exception:
                pass

    return {"ok": True, "deal_id": str(deal_id), "company_name": company_name}


def _deal_to_dict(d: Deal) -> dict:
    return {
        "id": str(d.id),
        "company_name": d.company_name,
        "bucket": d.bucket,
        "stage": d.stage,
        "location": d.location,
        "deal_size_m": float(d.deal_size_m) if d.deal_size_m is not None else None,
        "sector_primary": d.sector_primary,
        "sector_full": d.sector_full,
        "subsector": d.subsector,
        "security": d.security,
        "uop": d.uop,
        "source": d.source,
        "timing_qtr": d.timing_qtr,
        "nda": d.nda,
        "dataroom": d.dataroom,
        "mgmt_meeting": d.mgmt_meeting,
        "ioi_offered": d.ioi_offered,
        "ioi_signed": d.ioi_signed,
        "target_close": d.target_close.isoformat() if d.target_close else None,
        "commentary": d.commentary,
        "reasons_for_passing": d.reasons_for_passing,
        "last_updated": d.last_updated.isoformat() if d.last_updated else None,
        "updated_by": d.updated_by,
        "total_funded_m": float(d.total_funded_m) if d.total_funded_m is not None else None,
        # Pipeline model
        "pipeline_stage": d.pipeline_stage,
        "status": d.status,
        # Structural / financial / covenant fields
        "state": d.state,
        "next_action": d.next_action,
        "sourcing_date": d.sourcing_date.isoformat() if d.sourcing_date else None,
        "contact_name": d.contact_name,
        "contact_role": d.contact_role,
        "nda_date": d.nda_date.isoformat() if d.nda_date else None,
        "nda_status": d.nda_status,
        "tenor_months": d.tenor_months,
        "amortization": d.amortization,
        "oid_pct": float(d.oid_pct) if d.oid_pct is not None else None,
        "sofr_floor_pct": float(d.sofr_floor_pct) if d.sofr_floor_pct is not None else None,
        "call_protection": d.call_protection,
        "maturity_date": d.maturity_date.isoformat() if d.maturity_date else None,
        "total_leverage": float(d.total_leverage) if d.total_leverage is not None else None,
        "spread_bps": d.spread_bps,
        "base_rate": d.base_rate,
        "sofr_rate": float(d.sofr_rate) if d.sofr_rate is not None else None,
        "all_in_rate": float(d.all_in_rate) if d.all_in_rate is not None else None,
        "hold_amount_m": float(d.hold_amount_m) if d.hold_amount_m is not None else None,
        "ltm_revenue_m": float(d.ltm_revenue_m) if d.ltm_revenue_m is not None else None,
        "ltm_ebitda_m": float(d.ltm_ebitda_m) if d.ltm_ebitda_m is not None else None,
        "ebitda_margin": float(d.ebitda_margin) if d.ebitda_margin is not None else None,
        "revenue_growth_pct": float(d.revenue_growth_pct) if d.revenue_growth_pct is not None else None,
        "ebitda_growth_pct": float(d.ebitda_growth_pct) if d.ebitda_growth_pct is not None else None,
        "capex_m": float(d.capex_m) if d.capex_m is not None else None,
        "fcf_m": float(d.fcf_m) if d.fcf_m is not None else None,
        "dscr": float(d.dscr) if d.dscr is not None else None,
        "fccr": float(d.fccr) if d.fccr is not None else None,
        "interest_coverage": float(d.interest_coverage) if d.interest_coverage is not None else None,
        "max_leverage_covenant": float(d.max_leverage_covenant) if d.max_leverage_covenant is not None else None,
        "min_fccr_covenant": float(d.min_fccr_covenant) if d.min_fccr_covenant is not None else None,
        "capex_limit_covenant_m": float(d.capex_limit_covenant_m) if d.capex_limit_covenant_m is not None else None,
        "employees": d.employees,
        "locations_count": d.locations_count,
        "year_founded": d.year_founded,
        "risk_score": float(d.risk_score) if d.risk_score is not None else None,
        "deal_team": d.deal_team,
        "company_id": str(d.company_id) if d.company_id else None,
        "underwriting_locked": is_underwriting_locked(d.pipeline_stage),
    }


@router.get("/deals")
async def list_deals(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).order_by(Deal.bucket, Deal.company_name))
    deals = result.scalars().all()
    return [_deal_to_dict(d) for d in deals]


@router.get("/deals/{deal_id}")
async def get_deal(deal_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")
    return _deal_to_dict(deal)


@router.get("/deals/{deal_id}/underwriting/export")
async def export_underwriting(deal_id: uuid.UUID, db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal).where(Deal.id == deal_id))
    deal = result.scalar_one_or_none()
    if not deal:
        raise HTTPException(status_code=404, detail="Deal not found")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Underwriting"

    ws["A1"], ws["B1"] = "Deal Size ($M)", float(deal.deal_size_m) if deal.deal_size_m is not None else None
    ws["A2"], ws["B2"] = "LTM EBITDA ($M)", float(deal.ltm_ebitda_m) if deal.ltm_ebitda_m is not None else None
    ws["A3"], ws["B3"] = "SOFR Rate (%)", float(deal.sofr_rate) if deal.sofr_rate is not None else None
    ws["A4"], ws["B4"] = "Spread (bps)", deal.spread_bps
    # Formula cells — mirrors create_deal's derivation exactly (including its
    # None/zero guards, since incomplete underwriting fields are common), but
    # as a live Excel formula so the workbook stays an auditable model, not a
    # snapshot. IF-guards keep an incomplete input blank instead of #DIV/0!
    # or silently treating a blank cell as zero.
    ws["A5"], ws["B5"] = "Total Leverage (x)", '=IF(OR(B2="",B2=0),"",B1/B2)'
    ws["A6"], ws["B6"] = "All-In Rate (%)", '=IF(OR(B3="",B4=""),"",B3+B4/100)'

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = re.sub(r"[^A-Za-z0-9_-]+", "_", deal.company_name).strip("_") or "deal"
    filename = f"{safe_name}_Underwriting.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


class CreateDealRequest(BaseModel):
    company_name: str
    location: Optional[str] = None
    sector_primary: Optional[str] = None
    sector_full: Optional[str] = None
    subsector: Optional[str] = None
    security: Optional[str] = None
    uop: Optional[str] = None
    source: Optional[str] = None
    pipeline_stage: str = "sourcing"
    status: str = "Active"
    state: Optional[str] = None
    contact_name: Optional[str] = None
    contact_role: Optional[str] = None
    employees: Optional[int] = None
    locations_count: Optional[int] = None
    year_founded: Optional[int] = None
    deal_size_m: Optional[float] = None
    hold_amount_m: Optional[float] = None
    tenor_months: Optional[int] = None
    oid_pct: Optional[float] = None
    spread_bps: Optional[int] = None
    base_rate: Optional[str] = "SOFR"
    sofr_rate: Optional[float] = None
    sofr_floor_pct: Optional[float] = None
    ltm_revenue_m: Optional[float] = None
    ltm_ebitda_m: Optional[float] = None
    capex_m: Optional[float] = None
    ebitda_margin: Optional[float] = None
    revenue_growth_pct: Optional[float] = None
    max_leverage_covenant: Optional[float] = None
    min_fccr_covenant: Optional[float] = None
    capex_limit_covenant_m: Optional[float] = None


@router.post("/deals")
async def create_deal(
    body: CreateDealRequest,
    db: AsyncSession = Depends(get_db),
    auth: dict = Depends(require_auth),
):
    if body.pipeline_stage not in PIPELINE_STAGES:
        raise HTTPException(status_code=400, detail=f"Invalid pipeline_stage: {body.pipeline_stage!r}")
    if body.status not in STATUSES:
        raise HTTPException(status_code=400, detail=f"Invalid status: {body.status!r}")

    # Derived fields computed the same way the mockup's New Deal form does,
    # so a freshly-created deal shows a sensible leverage/all-in-rate immediately.
    all_in_rate = None
    if body.sofr_rate is not None and body.spread_bps is not None:
        all_in_rate = round(body.sofr_rate + body.spread_bps / 100, 4)
    total_leverage = None
    if body.deal_size_m is not None and body.ltm_ebitda_m:
        total_leverage = round(body.deal_size_m / body.ltm_ebitda_m, 2)

    # Every deal gets a backing Company row (Corporate Credit Data Model
    # v0.2 normalizes the borrower out of the deal) — mirrors migration 020's
    # one-company-per-deal backfill so deals created after that migration
    # aren't left with a permanently null company_id.
    company = Company(
        company_name=body.company_name,
        state=body.state,
        hq_location=body.location,
        sector=body.sector_primary,
        subsector=body.subsector,
    )
    db.add(company)
    await db.flush()

    deal = Deal(
        company_id=company.company_id,
        company_name=body.company_name,
        location=body.location,
        sector_primary=body.sector_primary,
        sector_full=body.sector_full,
        subsector=body.subsector,
        security=body.security,
        uop=body.uop,
        source=body.source,
        bucket="Active-Discussions",
        stage="Initial Conversations",
        pipeline_stage=body.pipeline_stage,
        status=body.status,
        state=body.state,
        contact_name=body.contact_name,
        contact_role=body.contact_role,
        employees=body.employees,
        locations_count=body.locations_count,
        year_founded=body.year_founded,
        deal_size_m=body.deal_size_m,
        hold_amount_m=body.hold_amount_m,
        tenor_months=body.tenor_months,
        oid_pct=body.oid_pct,
        spread_bps=body.spread_bps,
        base_rate=body.base_rate,
        sofr_rate=body.sofr_rate,
        sofr_floor_pct=body.sofr_floor_pct,
        all_in_rate=all_in_rate,
        total_leverage=total_leverage,
        ltm_revenue_m=body.ltm_revenue_m,
        ltm_ebitda_m=body.ltm_ebitda_m,
        capex_m=body.capex_m,
        ebitda_margin=body.ebitda_margin,
        revenue_growth_pct=body.revenue_growth_pct,
        max_leverage_covenant=body.max_leverage_covenant,
        min_fccr_covenant=body.min_fccr_covenant,
        capex_limit_covenant_m=body.capex_limit_covenant_m,
        updated_by="manual_edit",
    )
    db.add(deal)
    await db.flush()

    if deal.pipeline_stage == "portfolio_monitoring":
        await ensure_portfolio_position(deal, db)

    db.add(DealUpdateLog(
        deal_id=deal.id,
        field_changed="pipeline_stage",
        old_value=None,
        new_value=deal.pipeline_stage,
        source="manual_edit",
    ))
    await log_activity(db, deal.id, get_actor_name(auth), "system", "Deal created — entered via New Deal form")

    return {"ok": True, "deal_id": str(deal.id), "company_name": deal.company_name}


@router.get("/kpis")
async def get_kpis(db: AsyncSession = Depends(get_db)):
    result = await db.execute(select(Deal))
    all_deals = result.scalars().all()
    closed = [d for d in all_deals if d.bucket == "Closed"]
    return {
        "total_reviewed": len(all_deals),
        "closed": len(closed),
        "active_diligence": sum(1 for d in all_deals if d.bucket == "Active-Diligence"),
        "active_discussions": sum(1 for d in all_deals if d.bucket == "Active-Discussions"),
        "passed": sum(1 for d in all_deals if d.bucket == "Dead-Hold"),
        "deployed_m": float(sum(d.total_funded_m or 0 for d in closed)),
    }


@router.get("/logs/deal-updates")
async def get_deal_update_logs(
    db: AsyncSession = Depends(get_db),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
    deal_id: Optional[uuid.UUID] = Query(None),
):
    q = (
        select(DealUpdateLog, Deal.company_name)
        .join(Deal, DealUpdateLog.deal_id == Deal.id)
        .order_by(desc(DealUpdateLog.changed_at))
        .limit(limit)
        .offset(offset)
    )
    if deal_id is not None:
        q = q.where(DealUpdateLog.deal_id == deal_id)
    rows = (await db.execute(q)).all()
    return [
        {
            "id": log.id,
            "deal_id": str(log.deal_id),
            "company_name": name,
            "field_changed": log.field_changed,
            "old_value": log.old_value,
            "new_value": log.new_value,
            "source": log.source,
            "changed_at": log.changed_at.isoformat(),
            "email_subject": log.email_subject,
        }
        for log, name in rows
    ]


@router.get("/logs/email-scans")
async def get_email_scan_logs(
    db: AsyncSession = Depends(get_db),
    limit: int = Query(100, ge=1, le=500),
    offset: int = Query(0, ge=0),
):
    q = (
        select(EmailScanLog, Deal.company_name)
        .outerjoin(Deal, EmailScanLog.matched_deal_id == Deal.id)
        .order_by(desc(EmailScanLog.processed_at))
        .limit(limit)
        .offset(offset)
    )
    rows = (await db.execute(q)).all()
    return [
        {
            "id": log.id,
            "subject": log.subject,
            "user_email": log.user_email,
            "received_at": log.received_at.isoformat() if log.received_at else None,
            "processed_at": log.processed_at.isoformat(),
            "matched_deal_id": str(log.matched_deal_id) if log.matched_deal_id else None,
            "company_name": name,
            "claude_summary": log.claude_summary,
            "action_taken": log.action_taken,
        }
        for log, name in rows
    ]


@router.get("/analytics")
async def get_analytics(db: AsyncSession = Depends(get_db)):
    all_deals = (await db.execute(select(Deal))).scalars().all()
    funnel = {
        "total_reviewed": len(all_deals),
        "nda_signed": sum(1 for d in all_deals if d.nda == "P"),
        "closed": sum(1 for d in all_deals if d.bucket == "Closed"),
    }

    pass_rows = await db.execute(
        text(
            "SELECT reasons_for_passing, COUNT(*) AS cnt FROM credit_deals "
            "WHERE bucket='Dead-Hold' AND reasons_for_passing IS NOT NULL AND reasons_for_passing != '' "
            "GROUP BY reasons_for_passing ORDER BY cnt DESC"
        )
    )
    pass_reasons = [{"reason": row[0], "count": int(row[1])} for row in pass_rows]

    source_rows = await db.execute(
        text(
            "SELECT source, COUNT(*) AS cnt FROM credit_deals "
            "WHERE source IS NOT NULL AND source != '' "
            "GROUP BY source ORDER BY cnt DESC LIMIT 12"
        )
    )
    deal_sources = [{"source": row[0], "count": int(row[1])} for row in source_rows]

    quarter_rows = await db.execute(
        text(
            "SELECT timing_qtr, COUNT(*) AS cnt FROM credit_deals "
            "WHERE timing_qtr IS NOT NULL AND timing_qtr != '' "
            "GROUP BY timing_qtr ORDER BY SUBSTRING(timing_qtr FROM 4), LEFT(timing_qtr, 1)"
        )
    )
    deals_by_quarter = [{"quarter": row[0], "count": int(row[1])} for row in quarter_rows]

    return {
        "funnel": funnel,
        "pass_reasons": pass_reasons,
        "deal_sources": deal_sources,
        "deals_by_quarter": deals_by_quarter,
    }
